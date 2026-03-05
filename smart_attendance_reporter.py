import customtkinter as ctk
from tkinter import messagebox
from tkcalendar import DateEntry
from zk import ZK
from datetime import datetime, timedelta
import pandas as pd
import os
from collections import defaultdict
from openpyxl.styles import Font, PatternFill

# ================= DEVICE =================
DEVICE_IP = "192.168.6.251"
# ==========================================

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class AttendanceApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Smart Attendance Reporter PRO - By Islam Abdeen")
        self.geometry("750x720")
        self.resizable(False, False)

        # Center Window
        self.update_idletasks()
        width = 750
        height = 720
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        title = ctk.CTkLabel(self, text="SMART ATTENDANCE REPORTER PRO",
                             font=ctk.CTkFont(size=22, weight="bold"))
        title.pack(pady=15)

        subtitle = ctk.CTkLabel(self, text="Licensed Version - By Islam Abdeen",
                                font=ctk.CTkFont(size=14))
        subtitle.pack()

        main_frame = ctk.CTkFrame(self)
        main_frame.pack(pady=20, padx=40, fill="both", expand=True)

        # ================= Date From =================
        ctk.CTkLabel(main_frame, text="Date From").pack(pady=5)
        self.from_date = DateEntry(main_frame, date_pattern='dd-mm-yyyy')
        self.from_date.pack()

        tf1 = ctk.CTkFrame(main_frame)
        tf1.pack(pady=5)

        self.from_hour = ctk.CTkComboBox(tf1, values=[f"{h:02d}" for h in range(24)], width=80)
        self.from_hour.set("00")
        self.from_hour.pack(side="left", padx=5)

        self.from_minute = ctk.CTkComboBox(tf1, values=[f"{m:02d}" for m in range(60)], width=80)
        self.from_minute.set("00")
        self.from_minute.pack(side="left", padx=5)

        # ================= Date To =================
        ctk.CTkLabel(main_frame, text="Date To").pack(pady=5)
        self.to_date = DateEntry(main_frame, date_pattern='dd-mm-yyyy')
        self.to_date.pack()

        tf2 = ctk.CTkFrame(main_frame)
        tf2.pack(pady=5)

        self.to_hour = ctk.CTkComboBox(tf2, values=[f"{h:02d}" for h in range(24)], width=80)
        self.to_hour.set("23")
        self.to_hour.pack(side="left", padx=5)

        self.to_minute = ctk.CTkComboBox(tf2, values=[f"{m:02d}" for m in range(60)], width=80)
        self.to_minute.set("59")
        self.to_minute.pack(side="left", padx=5)

        # ================= Employee =================
        ctk.CTkLabel(main_frame, text="Employee").pack(pady=10)
        self.employee_combo = ctk.CTkComboBox(main_frame, values=["All"], width=250)
        self.employee_combo.set("All")
        self.employee_combo.pack()

        # ================= Buttons =================
        self.generate_btn = ctk.CTkButton(main_frame,
                                          text="Generate Excel Report",
                                          height=40,
                                          font=ctk.CTkFont(size=15, weight="bold"),
                                          command=self.generate_excel)
        self.generate_btn.pack(pady=25)

        self.status_label = ctk.CTkLabel(main_frame, text="", text_color="green")
        self.status_label.pack(pady=5)

        # Load Employees Automatically
        self.after(500, self.load_employees)

    # ================= Load Employees =================
    def load_employees(self):
        try:
            zk = ZK(DEVICE_IP, port=4370, timeout=10)
            conn = zk.connect()
            users = conn.get_users()
            conn.disconnect()

            values = ["All"] + [str(u.user_id) for u in users]
            self.employee_combo.configure(values=values)
            self.employee_combo.set("All")

        except:
            pass

    # ================= Generate Report =================
    def generate_excel(self):

        self.status_label.configure(text="Processing... Please wait")
        self.update()

        try:
            start_datetime = datetime.strptime(
                f"{self.from_date.get()} {self.from_hour.get()}:{self.from_minute.get()}",
                "%d-%m-%Y %H:%M"
            )
            end_datetime = datetime.strptime(
                f"{self.to_date.get()} {self.to_hour.get()}:{self.to_minute.get()}",
                "%d-%m-%Y %H:%M"
            )
        except:
            messagebox.showerror("Error", "Invalid Date Format")
            return

        try:
            zk = ZK(DEVICE_IP, port=4370, timeout=20)
            conn = zk.connect()
            attendance = conn.get_attendance()
            users = conn.get_users()
            conn.disconnect()
        except Exception as e:
            messagebox.showerror("Device Error ❌", str(e))
            return

        user_dict = {str(u.user_id): u.name for u in users}
        selected_employee = self.employee_combo.get()

        records = []
        for rec in attendance:
            if start_datetime <= rec.timestamp <= end_datetime:
                if selected_employee != "All" and str(rec.user_id) != selected_employee:
                    continue

                records.append({
                    "user_id": str(rec.user_id),
                    "timestamp": rec.timestamp,
                    "punch": rec.punch
                })

        records.sort(key=lambda x: (x["user_id"], x["timestamp"]))

        grouped = defaultdict(list)
        for r in records:
            grouped[r["user_id"]].append(r)

        final_rows = []

        for user_id, logs in grouped.items():
            open_checkin = None

            for log in logs:
                ts = log["timestamp"]
                punch = log["punch"]

                if punch == 0:
                    if open_checkin is None:
                        open_checkin = ts
                    else:
                        final_rows.append([user_id, user_dict.get(user_id, ""), open_checkin, ""])
                        open_checkin = ts

                elif punch == 1:
                    if open_checkin:
                        if (ts - open_checkin) <= timedelta(hours=24):
                            final_rows.append([user_id, user_dict.get(user_id, ""), open_checkin, ts])
                            open_checkin = None
                        else:
                            final_rows.append([user_id, user_dict.get(user_id, ""), open_checkin, ""])
                            final_rows.append([user_id, user_dict.get(user_id, ""), "", ts])
                            open_checkin = None
                    else:
                        final_rows.append([user_id, user_dict.get(user_id, ""), "", ts])

            if open_checkin:
                final_rows.append([user_id, user_dict.get(user_id, ""), open_checkin, ""])

        if not final_rows:
            messagebox.showinfo("Info", "No records found")
            return

        df = pd.DataFrame(final_rows,
                          columns=["Employee Code", "Employee Name", "Check-in", "Check-out"])

        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        base_name = "Smart_Attendance_Report_PRO"
        counter = 1

        while True:
            file_path = os.path.join(downloads, f"{base_name}_{counter}.xlsx")
            if not os.path.exists(file_path):
                break
            counter += 1

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Details", index=False)
            ws = writer.sheets["Details"]

            headers = ["Employee Code", "Employee Name", "Check-in", "Check-out", "Duration"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True)

            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            for row in range(2, ws.max_row + 1):

                ws[f"C{row}"].number_format = "DD-MM-YYYY HH:MM"
                ws[f"D{row}"].number_format = "DD-MM-YYYY HH:MM"

                ws[f"E{row}"] = f'=IF(AND(C{row}<>"",D{row}<>""),D{row}-C{row},"")'
                ws[f"E{row}"].number_format = "[h]:mm"

                if ws[f"C{row}"].value == "" or ws[f"D{row}"].value == "":
                    ws[f"C{row}"].fill = red_fill
                    ws[f"D{row}"].fill = red_fill

            # Auto column width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

        self.status_label.configure(text="Report Generated Successfully ✅")
        messagebox.showinfo("Success", "Report Saved in Downloads ✅")


if __name__ == "__main__":
    app = AttendanceApp()
    app.mainloop()